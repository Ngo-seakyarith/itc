from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_XLSX = Path("Engine_Oil_Interpolated.xlsx")

HEADERS = [
    ("T (K)", "0.0"),
    ("rho (kg/m^3)", "0.0"),
    ("cp (kJ/kg.K)", "0.000"),
    ("mu x 10^2 (N.s/m^2)", "0.00"),
    ("nu x 10^6 (m^2/s)", "0.0"),
    ("k x 10^3 (W/m.K)", "0.0"),
    ("alpha x 10^7 (m^2/s)", "0.000"),
    ("Pr", "#,##0"),
    ("beta x 10^3 (K^-1)", "0.00"),
]

ENGINE_OIL_DATA = [
    [273, 899.1, 1.796, 385.0, 4280.0, 147.0, 0.910, 47000, 0.70],
    [280, 895.3, 1.827, 217.0, 2430.0, 144.0, 0.880, 27500, 0.70],
    [290, 890.0, 1.868, 99.9, 1120.0, 145.0, 0.872, 12900, 0.70],
    [300, 884.1, 1.909, 48.6, 550.0, 145.0, 0.859, 6400, 0.70],
    [310, 877.9, 1.951, 25.3, 288.0, 145.0, 0.847, 3400, 0.70],
    [320, 871.8, 1.993, 14.1, 161.0, 143.0, 0.823, 1965, 0.70],
    [330, 865.8, 2.035, 8.36, 96.6, 141.0, 0.800, 1205, 0.70],
    [340, 859.9, 2.076, 5.31, 61.7, 139.0, 0.779, 793, 0.70],
    [350, 853.9, 2.118, 3.56, 41.7, 138.0, 0.763, 546, 0.70],
    [360, 847.8, 2.161, 2.52, 29.7, 138.0, 0.753, 395, 0.70],
    [370, 841.8, 2.206, 1.86, 22.0, 137.0, 0.738, 300, 0.70],
    [380, 836.0, 2.250, 1.41, 16.9, 136.0, 0.723, 233, 0.70],
    [390, 830.6, 2.294, 1.10, 13.3, 135.0, 0.709, 187, 0.70],
]


def original_excel_row(index: int) -> int:
    return 2 + index * 2


def interpolated_excel_row(index: int) -> int:
    return 3 + index * 2


def write_headers(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    ws["A1"] = "Point Type"
    ws["A1"].fill = header_fill
    ws["A1"].font = header_font

    for col_idx, (title, _) in enumerate(HEADERS, start=2):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font


def write_original_rows(ws):
    for data_index, row_data in enumerate(ENGINE_OIL_DATA):
        row_number = original_excel_row(data_index)
        ws.cell(row=row_number, column=1, value="Original")

        for col_idx, value in enumerate(row_data, start=2):
            ws.cell(row=row_number, column=col_idx, value=value)


def write_interpolated_rows(ws):
    red_font = Font(color="FF0000")
    red_fill = PatternFill(fill_type="solid", fgColor="FDE9E7")

    for data_index in range(len(ENGINE_OIL_DATA) - 1):
        row_number = interpolated_excel_row(data_index)
        prev_row = original_excel_row(data_index)
        next_row = original_excel_row(data_index + 1)

        label_cell = ws.cell(row=row_number, column=1, value="Interpolated")
        label_cell.font = red_font
        label_cell.fill = red_fill

        temperature_col = get_column_letter(2)
        temp_cell = ws.cell(row=row_number, column=2)
        temp_cell.value = f"=({temperature_col}{prev_row}+{temperature_col}{next_row})/2"
        temp_cell.font = red_font
        temp_cell.fill = red_fill

        for col_idx in range(3, len(HEADERS) + 2):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_number, column=col_idx)
            cell.value = (
                f"={col_letter}{prev_row}"
                f"+($B{row_number}-$B{prev_row})"
                f"*({col_letter}{next_row}-{col_letter}{prev_row})"
                f"/($B{next_row}-$B{prev_row})"
            )
            cell.font = red_font
            cell.fill = red_fill


def apply_number_formats(ws):
    for col_idx, (_, number_format) in enumerate(HEADERS, start=2):
        for row_idx in range(2, 2 * len(ENGINE_OIL_DATA) + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = number_format


def size_columns(ws):
    widths = {
        "A": 16,
        "B": 10,
        "C": 16,
        "D": 16,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 20,
        "I": 12,
        "J": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def build_workbook(output_path: Path):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Engine Oil Data"

    write_headers(worksheet)
    write_original_rows(worksheet)
    write_interpolated_rows(worksheet)
    apply_number_formats(worksheet)
    size_columns(worksheet)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:J{2 * len(ENGINE_OIL_DATA) - 1}"

    workbook.save(output_path)


def main():
    output_path = Path(__file__).with_name(OUTPUT_XLSX.name)
    build_workbook(output_path)
    print(f"Created {output_path}")


if __name__ == "__main__":
    main()
